

from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter



# PLAXIS EXCEL EXPORT SETTINGS
# =============================================================================

CREATE_PLAXIS_EXCEL = True

DEFAULT_BOREHOLE_NAME = "BH1"
DEFAULT_X_COORD = 0.0
DEFAULT_Y_COORD = 0.0

DEFAULT_MATERIAL_MODEL = "Linear"
DEFAULT_POISSON_RATIO = 0.30
DEFAULT_TENSILE_STRENGTH_KPA = 2.0
DEFAULT_INTERFACE_STRENGTH = 0.67
DEFAULT_K0 = "F"
DEFAULT_K0X_EQUALS_K0Y = "F"
DEFAULT_K0X = 1.0
DEFAULT_K0Y = 1.0

# Important setting! First observe the soil distribution before making changes

PLAXIS_BOTTOM_ELEVATION_NAP = -30.0
SIMPLIFY_SAND_LAYERS_AT_BOTTOM = True

# =============================================================================
# PLAXIS EXPORT SETTINGS
# =============================================================================

PLAXIS_SOIL_FAMILY_MAP = {
    "Peat": "Peat",
    "Organic clay": "Organic clay",
    "Soil, fine grain": "Clay",
    "Clay & silt": "Clay & silt",
    "Silty clay / clayey silt": "Silty clay",
    "Sand mixtures": "Silty sand",
    "Sand": "Sand",
    "Dense sand / gravelly sand": "Dense sand",
    "Unknown material": "Unknown",
}


def _infer_soil_type_from_soil_id(soil_id: str) -> str:
    """Infer original soil type from a soil_id string.

    Examples:
        Sand 3 -> Sand
        Sand 3 + Sand 4 -> Sand
        Clay & silt 1 -> Clay & silt
        Dense sand / gravelly sand 1 -> Dense sand / gravelly sand
    """
    soil_id = str(soil_id)

    first_part = soil_id.split("+")[0].strip()
    parts = first_part.rsplit(" ", 1)

    if len(parts) == 2 and parts[1].isdigit():
        return parts[0]

    return first_part


def _create_clean_plaxis_names(df: pd.DataFrame) -> pd.DataFrame:
    """Create clean sequential PLAXIS material names.

    Example:
        Clay & silt 1 + Clay & silt 2 -> Clay & silt 1
        Clay & silt 3                 -> Clay & silt 2
        Sand 3 + Sand 4               -> Sand 1
    """
    out = df.copy()

    counters: dict[str, int] = {}
    plaxis_names = []
    plaxis_families = []

    for soil_id in out["soil_id"]:
        inferred_soil_type = _infer_soil_type_from_soil_id(soil_id)
        family = PLAXIS_SOIL_FAMILY_MAP.get(inferred_soil_type, inferred_soil_type)

        counters[family] = counters.get(family, 0) + 1

        plaxis_name = f"{family} {counters[family]}"

        plaxis_families.append(family)
        plaxis_names.append(plaxis_name)

    out["plaxis_soil_family"] = plaxis_families
    out["plaxis_material_name"] = plaxis_names

    return out


def _style_workbook(wb: Workbook) -> None:
    """Apply simple clean formatting to the PLAXIS input workbook."""
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 32)

def _is_sand_like_soil_id(soil_id: str) -> bool:
    """Return True if a soil_id belongs to a sand-like CPT class."""
    soil_type = _infer_soil_type_from_soil_id(soil_id)

    return soil_type in {
        "Sand",
        "Sand mixtures",
        "Dense sand / gravelly sand",
    }


def _clip_layers_to_bottom_elevation(
    df: pd.DataFrame,
    bottom_elevation_m_nap: float,
) -> pd.DataFrame:
    """Clip the layer table at a given bottom elevation.

    Layers fully below the bottom elevation are removed.
    A layer crossing the bottom elevation is cut to that elevation.
    """
    clipped_rows = []

    df = df.sort_values("top_elevation_m_nap", ascending=False).reset_index(drop=True)

    for _, row in df.iterrows():
        top = float(row["top_elevation_m_nap"])
        bottom = float(row["bottom_elevation_m_nap"])

        # Layer is completely below the export bottom.
        if top <= bottom_elevation_m_nap:
            continue

        new_row = row.copy()

        # Layer crosses the export bottom. Clip it and stop after this layer.
        if bottom < bottom_elevation_m_nap:
            new_row["bottom_elevation_m_nap"] = bottom_elevation_m_nap
            clipped_rows.append(new_row)
            break

        clipped_rows.append(new_row)

    if not clipped_rows:
        raise ValueError(
            "No layers remain after clipping to PLAXIS bottom elevation "
            f"{bottom_elevation_m_nap} m NAP."
        )

    return pd.DataFrame(clipped_rows).reset_index(drop=True)


def _thickness_weighted_average(
    rows: pd.DataFrame,
    column: str,
) -> float | None:
    """Calculate thickness-weighted average for one numeric column."""
    if column not in rows.columns:
        return None

    values = pd.to_numeric(rows[column], errors="coerce")

    if values.isna().all():
        return None

    thickness = (
        rows["top_elevation_m_nap"].astype(float)
        - rows["bottom_elevation_m_nap"].astype(float)
    ).abs()

    valid = values.notna() & thickness.notna() & (thickness > 0)

    if not valid.any():
        return None

    return float((values[valid] * thickness[valid]).sum() / thickness[valid].sum())


def _combine_bottom_sand_stack(df: pd.DataFrame) -> pd.DataFrame:
    """Combine the contiguous sand-like stack at the bottom of the profile.

    Starting from the deepest exported layer, move upward while layers are
    sand-like. Combine that bottom sand stack into one layer.

    The combined layer keeps the soil_id of the uppermost sand-like layer.
    Numeric parameters are thickness-weighted averages.
    """
    if df.empty:
        return df.copy()

    df = df.sort_values("top_elevation_m_nap", ascending=False).reset_index(drop=True)

    # If the deepest layer is not sand-like, there is nothing to simplify.
    if not _is_sand_like_soil_id(df.iloc[-1]["soil_id"]):
        return df

    start_idx = len(df) - 1

    while start_idx > 0 and _is_sand_like_soil_id(df.iloc[start_idx - 1]["soil_id"]):
        start_idx -= 1

    sand_stack = df.iloc[start_idx:].copy()
    upper_part = df.iloc[:start_idx].copy()

    combined_row = sand_stack.iloc[0].copy()

    # Keep the name of the uppermost sand-like layer.
    exact_sand_rows = sand_stack[
        sand_stack["soil_id"].apply(
            lambda value: _infer_soil_type_from_soil_id(value) == "Sand"
        )
    ]

    if not exact_sand_rows.empty:
        # Use the uppermost exact Sand layer name.
        combined_row["soil_id"] = exact_sand_rows.iloc[0]["soil_id"]
    else:
        # Fallback: use the uppermost sand-like layer name.
        combined_row["soil_id"] = sand_stack.iloc[0]["soil_id"]

    combined_row["top_elevation_m_nap"] = float(sand_stack.iloc[0]["top_elevation_m_nap"])
    combined_row["bottom_elevation_m_nap"] = float(sand_stack.iloc[-1]["bottom_elevation_m_nap"])

    numeric_weighted_cols = [
        "qc_mpa_mean",
        "gamma_unsat_kn_m3",
        "gamma_sat_kn_m3",
        "E100_mpa",
        "c_prime_kpa",
        "phi_prime_deg",
        "dilatancy_deg",
    ]

    for col in numeric_weighted_cols:
        if col in sand_stack.columns:
            combined_row[col] = _thickness_weighted_average(sand_stack, col)

    # Keep first non-numeric metadata from the uppermost sand-like layer.
    if "interpolation_status" in sand_stack.columns:
        combined_row["interpolation_status"] = "bottom_sand_stack_weighted_average"

    combined = pd.concat(
        [upper_part, combined_row.to_frame().T],
        ignore_index=True,
    )

    return combined.reset_index(drop=True)


def _prepare_layers_for_plaxis_export(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare effective layers for PLAXIS export.

    Applies PLAXIS-specific modelling rules:
        1. clip to DEFAULT_PLAXIS_BOTTOM_ELEVATION_M_NAP
        2. optionally combine bottom sand-like stack
    """
    prepared = _clip_layers_to_bottom_elevation(
        df,
        bottom_elevation_m_nap=PLAXIS_BOTTOM_ELEVATION_NAP,
    )

    if SIMPLIFY_SAND_LAYERS_AT_BOTTOM:
        prepared = _combine_bottom_sand_stack(prepared)


    return prepared.reset_index(drop=True)


def create_plaxis_soil_input_excel(
    effective_layer_parameters_df: pd.DataFrame,
    output_path: str | Path,
) -> pd.DataFrame:
    """Create a clean PLAXIS soil input Excel workbook.

    Parameters
    ----------
    effective_layer_parameters_df:
        DataFrame from eff_layer_parameters.csv.
    output_path:
        Path to the .xlsx file to create.

    Returns
    -------
    plaxis_df:
        DataFrame with generated PLAXIS material names.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    required_columns = {
        "soil_id",
        "top_elevation_m_nap",
        "bottom_elevation_m_nap",
        "gamma_sat_kn_m3",
        "E100_mpa",
        "c_prime_kpa",
        "phi_prime_deg",
        "dilatancy_deg",
    }

    missing = required_columns - set(effective_layer_parameters_df.columns)

    if missing:
        raise ValueError(
            "Cannot create PLAXIS Excel. "
            f"Missing columns in effective_layer_parameters_df: {sorted(missing)}"
        )

    df = _prepare_layers_for_plaxis_export(effective_layer_parameters_df)
    df = _create_clean_plaxis_names(df)

    wb = Workbook()

    # Remove default sheet.
    default_ws = wb.active
    wb.remove(default_ws)

    # -------------------------------------------------------------------------
    # Sheet 1: OHE Ground Profile
    # -------------------------------------------------------------------------
    ws_profile = wb.create_sheet("OHE Ground Profile")

    top_level = float(df.loc[0, "top_elevation_m_nap"])

    ws_profile["A1"] = "Name"
    ws_profile["B1"] = DEFAULT_BOREHOLE_NAME

    ws_profile["A2"] = "x coord"
    ws_profile["B2"] = DEFAULT_X_COORD

    ws_profile["A3"] = "y coord"
    ws_profile["B3"] = DEFAULT_Y_COORD

    ws_profile["A4"] = "Top"
    ws_profile["B4"] = top_level

    start_row = 5

    for i, row in df.iterrows():
        excel_row = start_row + i
        ws_profile.cell(row=excel_row, column=1).value = row["plaxis_material_name"]
        ws_profile.cell(row=excel_row, column=2).value = float(row["bottom_elevation_m_nap"])

    # -------------------------------------------------------------------------
    # Sheet 2: Soil Properties
    # -------------------------------------------------------------------------
    ws_props = wb.create_sheet("Soil Properties")

    headers = [
        "Unit",
        "Name",
        "Material",
        "Unit weight (kN/m3)",
        "E' (kPa)",
        "v'",
        "c' (kPa)",
        "phi",
        "dilatancy",
        "Tensile strength (kPa)",
        "Interface strength",
        "K0",
        "K0x =K0y",
        "K0x",
        "K0y",
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws_props.cell(row=1, column=col_idx).value = header

    for i, row in df.iterrows():
        excel_row = i + 2

        material_name = row["plaxis_material_name"]

        gamma_sat = float(row["gamma_sat_kn_m3"])
        e100_kpa = float(row["E100_mpa"]) * 1000.0
        c_prime = float(row["c_prime_kpa"])
        phi_prime = float(row["phi_prime_deg"])

        dilatancy = row["dilatancy_deg"]
        if pd.isna(dilatancy):
            dilatancy = 0.0
        else:
            dilatancy = float(dilatancy)

        values = [
            material_name,              # Unit
            material_name,              # Name
            DEFAULT_MATERIAL_MODEL,                   # Material
            gamma_sat,                  # Unit weight
            e100_kpa,                   # E' in kPa
            DEFAULT_POISSON_RATIO,      # v'
            c_prime,                    # c'
            phi_prime,                  # phi
            dilatancy,                  # dilatancy
            DEFAULT_TENSILE_STRENGTH_KPA,   # tensile strength
            DEFAULT_INTERFACE_STRENGTH,     # interface strength
            DEFAULT_K0,                     # K0
            DEFAULT_K0X_EQUALS_K0Y,         # K0x = K0y
            DEFAULT_K0X,                    # K0x
            DEFAULT_K0Y,                            # K0y
        ]

        for col_idx, value in enumerate(values, start=1):
            ws_props.cell(row=excel_row, column=col_idx).value = value

    _style_workbook(wb)

    # Number formatting.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"

    wb.save(output_path)

    return df